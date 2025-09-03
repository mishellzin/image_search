import os
import shutil
from tkinter import filedialog
from openpyxl import load_workbook

def copy_images(list_xlsx, base_dir, destination="copied_images"):
    wb = load_workbook(list_xlsx)
    ws = wb.active
    
    names = [str(cell.value).strip() for cell in ws['A'] if cell.value]
    names_set = set(names)
    
    if not os.path.exists(destination):
        os.makedirs(destination)
    
    for root, _, files in os.walk(base_dir):
        for file in files:
            if file.lower().endswith(".jpg"):
                extensionless_name = os.path.splitext(file)[0]
                if extensionless_name in names_set:
                    root_dir = os.path.join(root, file)
                    destination_file = os.path.join(destination, file)
                    
                    if os.path.exists(destination_file):
                        base, ext = os.path.splitext(file)
                        counter = 1
                        while os.path.exists(destination_file):
                            destination_file = os.path.join(destination, f"{base}_{counter}{ext}")
                            counter += 1
                    
                    shutil.copy2(root_dir, destination_file)
                    print(f"Copying: {root_dir} -> {destination_file}")

if __name__ == "__main__":
    list_xlsx = filedialog.askopenfilename() 
    base_dir = filedialog.askdirectory()
    copy_images(list_xlsx, base_dir)
